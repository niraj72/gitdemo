import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]
    mainList=[]
    def formaterr(self,n):
        dict1=n.copy()
        self.mainList.append(dict1)

    def getTestData(self):
        Dict = {}
        co=0
        book = openpyxl.load_workbook("D:\\Prasana\\Pycharm\\homepage_data.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            li=self.mainList
          #  if Dict.__len__() != 0 :
          #      self.formaterr(Dict)
            if "Homepage_" in sheet.cell(row=i, column=1).value:

                for j in range(2, sheet.max_column + 1): # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                self.formaterr(Dict)
        print(self.mainList)
        return(self.mainList)
#HomePageData().getTestData()

